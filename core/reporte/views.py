from django.shortcuts import render
import os
from api import settings
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.views import View
from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML,CSS
from weasyprint.text.fonts import FontConfiguration
from openpyxl import Workbook
#from django.http.response import HttpResponse
from datetime import datetime
from ..asignacion.models import MetaMensualEP
from ..respuesta.models import EPRespuesta,AORespuesta
from ..respuesta.filters import RespuestasEjeFilter,RespuestasAreaFilter
from ..eje_prevencion.models import EjeTrabajo


class ReportEjeExcel(View):
    def get(self, request, *args, **kwargs):
        if self.request.user.is_superuser:
            ep_respuestas=EPRespuesta.objects.all()
        else:
            ep_respuestas=EPRespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasEjeFilter(request.GET,queryset=ep_respuestas)
        respuestas=filtro.qs
        size=4
        wb=Workbook()
        ws=wb.active
        ws['B2']='REPORTE DE RESPUESTAS EJE DE PREVENCION'
        ws.merge_cells('B2:G2')
        ws['B4']='Información General'
        ws.merge_cells('B4:M4')
        ws['N4']='Personas'
        ws.merge_cells('N4:W4')
        ws['X4']='Etnias'
        ws.merge_cells('X4:AA4')
        ws['B5']='Fecha de respuesta'
        ws['C5']='Delegación'
        ws['D5']='Priorizado'
        ws['E5']='No Priorizado'
        ws['F5']='Lugar Específico'
        ws['G5']='Plan/Orden'
        ws['H5']='Eje de Trabajo'
        ws['I5']='Producto'
        ws['J5']='Subproducto'
        ws['K5']='Observaciones'
        ws['L5']='Cantidad'
        ws['M5']='Total Personas'
        ws['N5']='Niños'
        ws['O5']='Niñas'
        ws['P5']='Adolecentes Masculinos'
        ws['Q5']='Adolecentes Femeninas'
        ws['R5']='Jovenes Masculinos'
        ws['S5']='Jovenes Femeninas'
        ws['T5']='Adultos Masculinos'
        ws['U5']='Adultas Femeninas'
        ws['V5']='Adultos Mayores Masculinos'
        ws['W5']='Adultas Mayores Femeninas'
        ws['X5']='Xincas'
        ws['Y5']='Garifunas'
        ws['Z5']='Mayas'
        ws['AA5']='Ladinos'
        cont=6
        for respuesta in respuestas:
            ws.cell(row=cont,column=2).value = respuesta.respondido
            ws.cell(row=cont,column=3).value = respuesta.delegacion.__str__()
            ws.cell(row=cont,column=4).value = respuesta.lugar_priorizado.__str__()
            ws.cell(row=cont,column=5).value = respuesta.lugar_no_priorizado.__str__()
            ws.cell(row=cont,column=6).value = respuesta.lugar_especifico.__str__()
            ws.cell(row=cont,column=7).value = respuesta.plan.__str__()
            ws.cell(row=cont,column=8).value = respuesta.eje.__str__()
            ws.cell(row=cont,column=9).value = respuesta.producto.__str__()
            ws.cell(row=cont,column=10).value = respuesta.subproducto.__str__()
            ws.cell(row=cont,column=11).value = respuesta.observaciones
            ws.cell(row=cont,column=12).value = respuesta.cantidad
            ws.cell(row=cont,column=13).value = respuesta.cantidad_personas
            ws.cell(row=cont,column=14).value = respuesta.ninios
            ws.cell(row=cont,column=15).value = respuesta.ninias
            ws.cell(row=cont,column=16).value = respuesta.adolecentes_masculinos
            ws.cell(row=cont,column=17).value = respuesta.adolecentes_femeninos
            ws.cell(row=cont,column=18).value = respuesta.jovenes_masculinos
            ws.cell(row=cont,column=19).value = respuesta.jovenes_femeninos
            ws.cell(row=cont,column=20).value = respuesta.adultos_masculinos
            ws.cell(row=cont,column=21).value = respuesta.adultos_femeninos
            ws.cell(row=cont,column=22).value = respuesta.adultos_mayores_masculinos
            ws.cell(row=cont,column=23).value = respuesta.adultos_mayores_femeninos
            ws.cell(row=cont,column=24).value = respuesta.xinca
            ws.cell(row=cont,column=25).value = respuesta.garifuna
            ws.cell(row=cont,column=26).value = respuesta.maya
            ws.cell(row=cont,column=27).value = respuesta.ladino
            cont+=1
        
        
        ws[f'B{cont}']='Totales'
        ws.merge_cells(f'B{cont}:K{cont}')
        ws.cell(row=cont,column=12).value = f'=SUM(L6:L{cont-1})'
        ws.cell(row=cont,column=13).value = f'=SUM(M6:M{cont-1})'
        ws.cell(row=cont,column=14).value = f'=SUM(N6:N{cont-1})'
        ws.cell(row=cont,column=15).value = f'=SUM(O6:O{cont-1})'
        ws.cell(row=cont,column=16).value = f'=SUM(P6:P{cont-1})'
        ws.cell(row=cont,column=17).value = f'=SUM(Q6:Q{cont-1})'
        ws.cell(row=cont,column=18).value = f'=SUM(R6:R{cont-1})'
        ws.cell(row=cont,column=19).value = f'=SUM(S6:S{cont-1})'
        ws.cell(row=cont,column=20).value = f'=SUM(T6:T{cont-1})'
        ws.cell(row=cont,column=21).value = f'=SUM(U6:U{cont-1})'
        ws.cell(row=cont,column=22).value = f'=SUM(V6:V{cont-1})'
        ws.cell(row=cont,column=23).value = f'=SUM(W6:W{cont-1})'
        ws.cell(row=cont,column=24).value = f'=SUM(X6:X{cont-1})'
        ws.cell(row=cont,column=25).value = f'=SUM(Y6:Y{cont-1})'
        ws.cell(row=cont,column=26).value = f'=SUM(Z6:Z{cont-1})'
        ws.cell(row=cont,column=27).value = f'=SUM(AA6:AA{cont-1})'
        filename='Reporte_Eje_de_Prevencion.xlsx'
        response=HttpResponse(content_type='application/ms-excel')
        content=f'attachment; filename = {filename}'
        response['Content-Disposition']=content
        wb.save(response)
        return response


class ReporteTemplateView(TemplateView):
    template_name='reporte.html'
    
    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)

class ReporteMetaEPTemplateView(TemplateView):
    template_name='reporte_meta_eje.html'
    
    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self,request):
        metas=MetaMensualEP.objects.all()
        ejes=EjeTrabajo.objects.all()
        delegaciones=[]
        delegacion_name=[]
        anios=[]
        for meta in metas:
            if meta.delegacion.nombre not in delegacion_name:
                delegacion_name.append(meta.delegacion.nombre)
                delegaciones.append({
                    'id':meta.delegacion.pk,
                    'nombre':meta.delegacion.nombre
                })
            if meta.asignado.year not in anios:
                anios.append(meta.asignado.year)
        return render(request,self.template_name,{'delegaciones':delegaciones,'anios':anios,'ejes':ejes})

class ReporteEjePDF(View):
    template_name='reporte_eje_prevencion.html'
    
    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self,request):
        if self.request.user.is_superuser:
            ep_respuestas=EPRespuesta.objects.all()
        else:
            ep_respuestas=EPRespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasEjeFilter(request.GET,queryset=ep_respuestas)
        respuestas=[]
        for respuesta in reversed(filtro.qs):
            respuestas.append(respuesta)
        data={
            'entity':respuestas,
            'logo':'{}{}'.format(settings.STATIC_URL,'img/logo_pnc.png')
        }
        try:
            template=get_template(self.template_name)
        except:
            pass
        html=template.render(data)
        css_url=os.path.join(settings.BASE_DIR,'static/bootstrap/css/bootstrap.min.css')
        pdf=HTML(string=html,base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
        
        return HttpResponse(pdf,content_type='application/pdf')

class ReporteGraficaEjePDF(TemplateView):
    template_name='grafica_eje.html'

    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self,request):
        if self.request.user.is_superuser:
            ep_respuestas=EPRespuesta.objects.all()
        else:
            ep_respuestas=EPRespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasEjeFilter(request.GET,queryset=ep_respuestas)
        respuestas=filtro.qs
   
        cantidad_personas=0
        ninios=0
        ninias=0
        adolecentes_masculinos=0
        adolecentes_femeninos=0
        jovenes_masculinos=0
        jovenes_femeninos=0
        adultos_masculinos=0
        adultos_femeninos=0
        adultos_mayores_masculinos=0
        adultos_mayores_femeninos=0
        xincas=0
        garifunas=0
        mayas=0
        ladinos=0
        for respuesta in respuestas:
            cantidad_personas+=respuesta.cantidad_personas
            ninios+=respuesta.ninios
            ninias+=respuesta.ninias
            jovenes_masculinos+=respuesta.jovenes_masculinos
            jovenes_femeninos+=respuesta.jovenes_femeninos
            adolecentes_masculinos+=respuesta.adolecentes_masculinos
            adolecentes_femeninos+=respuesta.adolecentes_femeninos
            adultos_masculinos+=respuesta.adultos_masculinos
            adultos_femeninos+=respuesta.adultos_femeninos
            adultos_mayores_masculinos+=respuesta.adultos_mayores_masculinos
            adultos_mayores_femeninos+=respuesta.adultos_mayores_femeninos
            xincas+=respuesta.xinca
            garifunas+=respuesta.garifuna
            mayas+=respuesta.maya
            ladinos+=respuesta.ladino
        size_respuestas=len(respuestas)

        return render(request,self.template_name,{
            'cantidad_personas':cantidad_personas,
            'ninios':ninios,
            'ninias':ninias,
            'adolecentes_masculinos':adolecentes_masculinos,
            'adolecentes_femeninos':adolecentes_femeninos,
            'jovenes_masculinos':jovenes_masculinos,
            'jovenes_femeninos':jovenes_femeninos,
            'adultos_masculinos':adultos_masculinos,
            'adultos_femeninos':adultos_femeninos,
            'adultos_mayores_masculinos':adultos_mayores_masculinos,
            'adultos_mayores_femeninos':adultos_mayores_femeninos,
            'xincas':xincas,
            'garifunas':garifunas,
            'mayas':mayas,
            'ladinos':ladinos,
            'size_respuestas':size_respuestas
        })

#--------------------------------------------------------------------------------------------

class ReporteAreaPDF(View):
    template_name='reporte_area_operativa.html'
    
    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self,request):
        if self.request.user.is_superuser:
            ao_respuestas=AORespuesta.objects.all()
        else:
            ao_respuestas=AORespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasAreaFilter(request.GET,queryset=ao_respuestas)
        respuestas=[]
        for respuesta in reversed(filtro.qs):
            respuestas.append(respuesta)
        
        data={
            'entity':respuestas,
            'logo':'{}{}'.format(settings.STATIC_URL,'img/logo_pnc.png')
        }
        try:
            template=get_template(self.template_name)
        except:
            pass
        html=template.render(data)
        css_url=os.path.join(settings.BASE_DIR,'static/bootstrap/css/bootstrap.min.css')
        pdf=HTML(string=html,base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
        
        return HttpResponse(pdf,content_type='application/pdf')


class ReporteGraficaAreaPDF(TemplateView):
    template_name='grafica_area.html'

    @method_decorator(login_required)
    def dispatch(self, request,*args,**kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self,request):
        if self.request.user.is_superuser:
            ao_respuestas=AORespuesta.objects.all()
        else:
            ao_respuestas=AORespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasAreaFilter(request.GET,queryset=ao_respuestas)
        respuestas=[]
        for respuesta in reversed(filtro.qs):
            respuestas.append(respuesta)
   
        total_identificados=0
        hombres_identificados=0
        mujeres_identificadas=0
        autos_identificados=0
        motos_identificadas=0
        armas_identificadas=0
        total_consignados=0
        hombres_consignados=0
        mujeres_consignadas=0
        autos_consignados=0
        motos_consignadas=0
        armas_consignadas=0
        total_conducidos=0
        hombres_conducidos=0
        mujeres_conducidas=0
        total_solventes=0
        hombres_solventes=0
        mujeres_solventes=0
        autos_solventes=0
        motos_solventes=0
        armas_solventes=0
        total_recuperados=0
        autos_recuperados=0
        motos_recuperados=0
        armas_recuperados=0
        hombres_recuperados=0
        mujeres_recuperadas=0
        menores_recuperados=0
        for respuesta in respuestas:
            total_identificados+=respuesta.total_identificados
            hombres_identificados+=respuesta.hombres_identificados
            mujeres_identificadas+=respuesta.mujeres_identificadas
            autos_identificados+=respuesta.autos_identificados
            motos_identificadas+=respuesta.motos_identificadas
            armas_identificadas+=respuesta.armas_identificadas
            total_consignados+=respuesta.total_consignados
            hombres_consignados+=respuesta.hombres_consignados
            mujeres_consignadas+=respuesta.mujeres_consignadas
            autos_consignados+=respuesta.autos_consignados
            motos_consignadas+=respuesta.motos_consignadas
            armas_consignadas+=respuesta.armas_consignadas
            total_conducidos+=respuesta.total_conducidos
            hombres_conducidos+=respuesta.hombres_conducidos
            mujeres_conducidas+=respuesta.mujeres_conducidas
            total_solventes+=respuesta.total_solventes
            hombres_solventes+=respuesta.hombres_solventes
            mujeres_solventes+=respuesta.mujeres_solventes
            autos_solventes+=respuesta.autos_solventes
            motos_solventes+=respuesta.motos_solventes
            armas_solventes+=respuesta.armas_solventes
            total_recuperados+=respuesta.total_recuperados
            autos_recuperados+=respuesta.autos_recuperados
            motos_recuperados+=respuesta.motos_recuperados
            armas_recuperados+=respuesta.armas_recuperados
            hombres_recuperados+=respuesta.hombres_recuperados
            mujeres_recuperadas+=respuesta.mujeres_recuperadas
            menores_recuperados+=respuesta.menores_recuperados
        size_respuestas=len(respuestas)

        return render(request,self.template_name,{
            'total_identificados':total_identificados,
            'hombres_identificados':hombres_identificados,
            'mujeres_identificadas':mujeres_identificadas,
            'autos_identificados':autos_identificados,
            'motos_identificadas':motos_identificadas,
            'armas_identificadas':armas_identificadas,
            'total_consignados':total_consignados,
            'hombres_consignados':hombres_consignados,
            'mujeres_consignadas':mujeres_consignadas,
            'autos_consignados':autos_consignados,
            'motos_consignadas':motos_consignadas,
            'armas_consignadas':armas_consignadas,
            'total_conducidos':total_conducidos,
            'hombres_conducidos':hombres_conducidos,
            'mujeres_conducidas':mujeres_conducidas,
            'total_solventes':total_solventes,
            'hombres_solventes':hombres_solventes,
            'mujeres_solventes':mujeres_solventes,
            'autos_solventes':autos_solventes,
            'motos_solventes':motos_solventes,
            'armas_solventes':armas_solventes,
            'total_recuperados':total_recuperados,
            'autos_recuperados':autos_recuperados,
            'motos_recuperados':motos_recuperados,
            'armas_recuperados':armas_recuperados,
            'hombres_recuperados':hombres_recuperados,
            'mujeres_recuperadas':mujeres_recuperadas,
            'menores_recuperados':menores_recuperados,
            'size_respuestas':size_respuestas
        })


class ReportAreaExcel(View):
    def get(self, request, *args, **kwargs):
        if self.request.user.is_superuser:
            ep_respuestas=AORespuesta.objects.all()
        else:
            ep_respuestas=AORespuesta.objects.filter(usuario=self.request.user)

        filtro=RespuestasAreaFilter(request.GET,queryset=ep_respuestas)
        respuestas=filtro.qs
        size=len(respuestas)+4
        celda_total=f'B{size}'
        wb=Workbook()
        ws=wb.active
        ws['B2']='REPORTE DE RESPUESTAS AREA OPERATIVA'
        ws.merge_cells('B2:E2')
        ws[celda_total]='Totales'
        ws.merge_cells(f'B{celda_total}:I{celda_total}')
        ws['B4']='Información General'
        ws.merge_cells(f'B4:J4')
        ws['K4']='Identificados'
        ws.merge_cells(f'K4:P4')
        ws['Q4']='Consignados'
        ws.merge_cells(f'Q4:V4')
        ws['W4']='Conducidos'
        ws.merge_cells(f'W4:Y4')
        ws['Z4']='Solventes'
        ws.merge_cells(f'Z4:AE4')
        ws['AF4']='Recuperados'
        ws.merge_cells(f'AF4:AL4')
        ws['B5']='Fecha de respuesta'
        ws['C5']='Delegación'
        ws['D5']='Priorizado'
        ws['E5']='No Priorizado'
        ws['F5']='Lugar de Apoyo'
        ws['G5']='Plan/Orden'
        ws['H5']='Tipo de Operativo'
        ws['I5']='Observaciones'
        ws['J5']='Cantidad'
        ws['K5']='Hombres'
        ws['L5']='Mujeres'
        ws['M5']='Autos'
        ws['N5']='Motos'
        ws['O5']='Armas'
        ws['P5']='Total'
        ws['Q5']='Hombres'
        ws['R5']='Mujeres'
        ws['S5']='Autos'
        ws['T5']='Motos'
        ws['U5']='Armas'
        ws['V5']='Total'
        ws['W5']='Hombres'
        ws['X5']='Mujeres'
        ws['Y5']='Total'
        ws['Z5']='Hombres'
        ws['AA5']='Mujeres'
        ws['AB5']='Autos'
        ws['AC5']='Motos'
        ws['AD5']='Armas'
        ws['AE5']='Total'
        ws['AF5']='Hombres'
        ws['AG5']='Mujeres'
        ws['AH5']='Menores'
        ws['AI5']='Autos'
        ws['AJ5']='Motos'
        ws['AK5']='Armas'
        ws['AL5']='Total'
        
        cont=6
        
        for respuesta in respuestas:
            ws.cell(row=cont,column=2).value = respuesta.respondido
            ws.cell(row=cont,column=3).value = respuesta.delegacion.__str__()
            ws.cell(row=cont,column=4).value = respuesta.lugar_priorizado.__str__()
            ws.cell(row=cont,column=5).value = respuesta.lugar_no_priorizado.__str__()
            ws.cell(row=cont,column=6).value = respuesta.lugar_apoyo.__str__()
            ws.cell(row=cont,column=7).value = respuesta.plan.__str__()
            ws.cell(row=cont,column=8).value = respuesta.operativo.__str__()
            ws.cell(row=cont,column=9).value = respuesta.observaciones
            ws.cell(row=cont,column=10).value = respuesta.cantidad
            ws.cell(row=cont,column=11).value = respuesta.hombres_identificados
            ws.cell(row=cont,column=12).value = respuesta.mujeres_identificadas
            ws.cell(row=cont,column=13).value = respuesta.autos_identificados
            ws.cell(row=cont,column=14).value = respuesta.motos_identificadas
            ws.cell(row=cont,column=15).value = respuesta.armas_identificadas
            ws.cell(row=cont,column=16).value = respuesta.total_identificados
            ws.cell(row=cont,column=17).value = respuesta.hombres_consignados
            ws.cell(row=cont,column=18).value = respuesta.mujeres_consignadas
            ws.cell(row=cont,column=19).value = respuesta.autos_consignados
            ws.cell(row=cont,column=20).value = respuesta.motos_consignadas
            ws.cell(row=cont,column=21).value = respuesta.armas_consignadas
            ws.cell(row=cont,column=22).value = respuesta.total_consignados
            ws.cell(row=cont,column=23).value = respuesta.hombres_conducidos
            ws.cell(row=cont,column=24).value = respuesta.mujeres_conducidas
            ws.cell(row=cont,column=25).value = respuesta.total_conducidos
            ws.cell(row=cont,column=26).value = respuesta.hombres_solventes
            ws.cell(row=cont,column=27).value = respuesta.mujeres_solventes
            ws.cell(row=cont,column=28).value = respuesta.autos_solventes
            ws.cell(row=cont,column=29).value = respuesta.motos_solventes
            ws.cell(row=cont,column=30).value = respuesta.armas_solventes
            ws.cell(row=cont,column=31).value = respuesta.total_solventes
            ws.cell(row=cont,column=32).value = respuesta.hombres_recuperados
            ws.cell(row=cont,column=33).value = respuesta.mujeres_recuperadas
            ws.cell(row=cont,column=34).value = respuesta.menores_recuperados
            ws.cell(row=cont,column=35).value = respuesta.autos_recuperados
            ws.cell(row=cont,column=36).value = respuesta.motos_recuperados
            ws.cell(row=cont,column=37).value = respuesta.armas_recuperados
            ws.cell(row=cont,column=38).value = respuesta.total_recuperados
            cont+=1

        ws[f'B{cont}']='Totales'
        ws.merge_cells(f'B{cont}:I{cont}')
        ws.cell(row=cont,column=10).value = f'=SUM(J6:J{cont-1})'
        ws.cell(row=cont,column=11).value = f'=SUM(K6:K{cont-1})'
        ws.cell(row=cont,column=12).value = f'=SUM(L6:L{cont-1})'
        ws.cell(row=cont,column=13).value = f'=SUM(M6:M{cont-1})'
        ws.cell(row=cont,column=14).value = f'=SUM(N6:N{cont-1})'
        ws.cell(row=cont,column=15).value = f'=SUM(O6:O{cont-1})'
        ws.cell(row=cont,column=16).value = f'=SUM(P6:P{cont-1})'
        ws.cell(row=cont,column=17).value = f'=SUM(Q6:Q{cont-1})'
        ws.cell(row=cont,column=18).value = f'=SUM(R6:R{cont-1})'
        ws.cell(row=cont,column=19).value = f'=SUM(S6:S{cont-1})'
        ws.cell(row=cont,column=20).value = f'=SUM(T6:T{cont-1})'
        ws.cell(row=cont,column=21).value = f'=SUM(U6:U{cont-1})'
        ws.cell(row=cont,column=22).value = f'=SUM(V6:V{cont-1})'
        ws.cell(row=cont,column=23).value = f'=SUM(W6:W{cont-1})'
        ws.cell(row=cont,column=24).value = f'=SUM(X6:X{cont-1})'
        ws.cell(row=cont,column=25).value = f'=SUM(Y6:Y{cont-1})'
        ws.cell(row=cont,column=26).value = f'=SUM(Z6:Z{cont-1})'
        ws.cell(row=cont,column=27).value = f'=SUM(AA6:AA{cont-1})'
        ws.cell(row=cont,column=28).value = f'=SUM(AB6:AB{cont-1})'
        ws.cell(row=cont,column=29).value = f'=SUM(AC6:AC{cont-1})'
        ws.cell(row=cont,column=30).value = f'=SUM(AD6:AD{cont-1})'
        ws.cell(row=cont,column=31).value = f'=SUM(AE6:AE{cont-1})'
        ws.cell(row=cont,column=32).value = f'=SUM(AF6:AF{cont-1})'
        ws.cell(row=cont,column=33).value = f'=SUM(AG6:AG{cont-1})'
        ws.cell(row=cont,column=34).value = f'=SUM(AH6:AH{cont-1})'
        ws.cell(row=cont,column=35).value = f'=SUM(AI6:AI{cont-1})'
        ws.cell(row=cont,column=36).value = f'=SUM(AJ6:AJ{cont-1})'
        ws.cell(row=cont,column=37).value = f'=SUM(AK6:AK{cont-1})'
        ws.cell(row=cont,column=38).value = f'=SUM(AL6:AL{cont-1})'

        filename='Reporte_Area_Operativa.xlsx'
        response=HttpResponse(content_type='application/ms-excel')
        content=f'attachment; filename = {filename}'
        response['Content-Disposition']=content
        wb.save(response)
        return response